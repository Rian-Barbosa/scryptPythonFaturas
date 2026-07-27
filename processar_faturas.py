import os
import re
import fitz  # PyMuPDF
import openpyxl

def clean_value(val_str):
    """Clean PDF currency string to float (e.g. 'R$3.004,30 BRL' -> 3004.30)"""
    if not val_str:
        return 0.0
    val_str = val_str.replace("R$", "").replace("BRL", "").strip()
    val_str = val_str.replace(".", "").replace(",", ".")
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def parse_invoices_from_pdfs(pdf_paths):
    """Parse all invoices from a list of compiled PDF files"""
    invoice_db = {}
    
    for pdf_path in pdf_paths:
        if not os.path.exists(pdf_path):
            print(f"Aviso: Arquivo PDF não encontrado: {pdf_path}")
            continue
            
        print(f"Lendo PDF: {os.path.basename(pdf_path)}...")
        doc = fitz.open(pdf_path)
        current_invoice = None
        
        for page in doc:
            text = page.get_text()
            if not text:
                continue
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # Check if a new invoice starts on this page
            is_new_invoice = False
            for line in lines[:5]:
                if "Recibo para Monte Carlo" in line:
                    is_new_invoice = True
                    break
            
            if is_new_invoice:
                if current_invoice and current_invoice["tx_id"]:
                    invoice_db[current_invoice["tx_id"]] = current_invoice
                current_invoice = {
                    "tx_id": None,
                    "subtotal": 0.0,
                    "awareness": 0.0,
                    "performance": 0.0,
                    "post": 0.0,
                    "campaigns": []
                }
                
            if current_invoice is None:
                continue
                
            # Extract Transaction ID and Subtotal
            for idx, line in enumerate(lines):
                if "Identificação da transação" in line or "Identificacao da transacao" in line:
                    if idx + 1 < len(lines):
                        current_invoice["tx_id"] = lines[idx+1]
                elif "Subtotal:" in line:
                    match = re.search(r"Subtotal:\s*(R\$\s*[\d\.,]+\s*BRL|R\$\s*[\d\.,]+|[\d\.,]+)", line)
                    if match:
                        current_invoice["subtotal"] = clean_value(match.group(1))
            
            # Extract Campaigns and Posts
            for idx, line in enumerate(lines):
                # Date pattern: "De [dia] de [mês] de [ano]... a [dia] de [mês] de [ano]..."
                if re.search(r"^De\s+\d+\s+de\s+\w+\s+de\s+\d+.*a\s+\d+\s+de\s+\w+\s+de\s+\d+", line):
                    campaign_name = lines[idx-1] if idx - 1 >= 0 else ""
                    val_str = lines[idx+1] if idx + 1 < len(lines) else ""
                    try:
                        val = clean_value(val_str)
                        current_invoice["campaigns"].append((campaign_name, val))
                    except Exception:
                        pass
                        
        if current_invoice and current_invoice["tx_id"]:
            invoice_db[current_invoice["tx_id"]] = current_invoice
            
    # Aggregate values into categories
    for tx_id, inv in invoice_db.items():
        for name, val in inv["campaigns"]:
            name_upper = name.upper()
            if "AWARENESS" in name_upper or "LEADS" in name_upper:
                inv["awareness"] += val
            elif "PERFORMANCE" in name_upper:
                inv["performance"] += val
            elif "POST" in name_upper or name_upper.startswith("POST DO INSTAGRAM"):
                inv["post"] += val
            else:
                print(f"Campanha não categorizada em {tx_id}: '{name}' = R${val:.2f}")
                
    return invoice_db

def fill_excel_template(template_path, output_path, invoice_db):
    """Fill the Excel template with parsed invoice data using openpyxl"""
    print(f"Abrindo planilha modelo: {os.path.basename(template_path)}...")
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # We locate header row (usually row 1) and determine column indexes
    headers = [cell.value for cell in sheet[1]]
    
    try:
        col_tx_id = headers.index("ID da transação") + 1
        col_subtotal = headers.index("Fatura sem Imposto") + 1
        col_awareness = headers.index("Awareness") + 1
        col_performance = headers.index("Performance") + 1
        col_post = headers.index("Post") + 1
    except ValueError as e:
        print(f"Erro: Colunas necessárias não encontradas no cabeçalho. {e}")
        return
        
    filled_count = 0
    # Loop through all rows starting from row 2
    for r in range(2, sheet.max_row + 1):
        tx_id = sheet.cell(row=r, column=col_tx_id).value
        
        # Skip empty rows or totals row
        if not tx_id or "Valor total cobrado" in str(tx_id):
            continue
            
        tx_id_str = str(tx_id).strip()
        if tx_id_str in invoice_db:
            inv = invoice_db[tx_id_str]
            # Write values to the columns
            sheet.cell(row=r, column=col_subtotal, value=round(inv["subtotal"], 2))
            sheet.cell(row=r, column=col_awareness, value=round(inv["awareness"], 2))
            sheet.cell(row=r, column=col_performance, value=round(inv["performance"], 2))
            sheet.cell(row=r, column=col_post, value=round(inv["post"], 2))
            filled_count += 1
            
    wb.save(output_path)
    print(f"Processo concluído! Planilha salva em: {os.path.basename(output_path)}")
    print(f"Total de faturas preenchidas: {filled_count}")

if __name__ == "__main__":
    # Define the directory of the script and the 'pdfs' folder path
    current_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in locals() else "."
    pdfs_dir = os.path.join(current_dir, "pdfs")
    
    # Create the 'pdfs' directory if it doesn't exist
    if not os.path.exists(pdfs_dir):
        os.makedirs(pdfs_dir)
        print("Pasta 'pdfs' criada automaticamente.")
        print("Por favor, coloque os arquivos PDF das faturas dentro dela.")
        input("\nPressione Enter para fechar...")
        exit()
        
    # Scan the 'pdfs' folder for PDF files
    all_files = os.listdir(pdfs_dir)
    pdf_filenames = [
        f for f in all_files 
        if f.lower().endswith(".pdf") and "extrair dados de pdf" not in f.lower()
    ]
    
    if not pdf_filenames:
        print("Erro: Nenhum arquivo PDF de fatura encontrado na pasta 'pdfs'.")
        print("Por favor, insira os PDFs das faturas na pasta 'pdfs' e tente novamente.")
        input("\nPressione Enter para fechar...")
        exit()
        
    # Build complete paths for the PDFs
    pdfs = [os.path.join(pdfs_dir, f) for f in pdf_filenames]
    print(f"Arquivos PDF detectados na pasta 'pdfs': {pdf_filenames}")
    
    # Excel paths (remain in the main script folder)
    template = os.path.join(current_dir, "Fatura detalhada para preencher.xlsx")
    output = os.path.join(current_dir, "Fatura detalhada preenchida resultado.xlsx")
    
    if not os.path.exists(template):
        print(f"Erro: Planilha modelo '{template}' nao encontrada na pasta principal.")
        input("\nPressione Enter para fechar...")
        exit()
        
    # Run pipeline
    db = parse_invoices_from_pdfs(pdfs)
    fill_excel_template(template, output, db)


